import requests
import time
from bs4 import BeautifulSoup
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from io import BytesIO

def scrape_product_info(url, session, headers, wb, index):
    # Send a GET request to the URL and include the session cookies
    response = session.get(url, headers=headers)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        soup = BeautifulSoup(response.content, 'html.parser')
    
        # Find all product containers
        product_containers = soup.find('div', class_='product-info-main')
        product_infos = product_containers.find('div', class_='product-info')
        product_gallery = product_containers.find('div', class_='product-gallery')
        product_cross_ref = product_containers.find('div', id='cross-reference')

        # Get the active worksheet
        ws = wb.active

        # Headers for each column
        headers = ['Automann', 'Product Part Number', 'Product Name', 'Product Note', 'Product Image']

        # Write headers if the first row is empty
        if ws['A1'].value is None:
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        img_row = ws.max_row + 1
        product_info_rows = []

        product_info_table = product_cross_ref.find('div', class_='product-info-content').find('table')

        if product_info_table is None:
            print(index, '===>', 'product table is none')
        else: 
            print(index, '===>', 'product table is exist')
            product_automann = product_infos.find('div', class_='product-info-attributes').find_all('dl', class_='attribute')[1].find('dd', class_='value').text.strip()
            product_info_rows = product_info_table.find('tbody').find_all('tr')
            
            for i, product_info_row in enumerate(product_info_rows):
                product_part_number = product_info_row.find_all('td')[1].text.strip()
                product_name = product_info_row.find_all('td')[0].text.strip()
                product_note = product_info_row.find_all('td')[2].text.strip()

                # Get the next available row
                row = ws.max_row + 1

                # Write other product information to adjacent cells
                ws[f'A{row}'] = product_automann
                ws[f'B{row}'] = product_part_number
                ws[f'C{row}'] = product_name
                ws[f'D{row}'] = product_note

                print(f"Product information for {product_name} added to the Excel file.")

                if i == 0:  # Add image only for the first row of each item
                    product_image = product_gallery.find_all('img')[0]['src']

                    # Save the image data
                    image_response = requests.get(product_image)
                    image_data = BytesIO(image_response.content)

                    # Add image to Excel and resize the cell to fit the image
                    img = XLImage(image_data)
                    img.width = 100  # Adjust the width as needed
                    img.height = 100  # Adjust the height as needed
                    ws.add_image(img, f'E{row}')
                    img_cell = ws[f'E{row}']

                    # Adjust image alignment
                    img_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Calculate the number of rows for the current item and adjust row height
                num_rows_current_item = len(product_info_rows)
                if num_rows_current_item < 6:
                    ws.row_dimensions[row].height = 100 / num_rows_current_item

            # Merge cells in column A and E per item
            start_row = ws.max_row - len(product_info_rows) + 1
            end_row = ws.max_row
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)  # Merge cells in column A
            ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)  # Merge cells in column E

            # Resize the column width to fit the content
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 20

            # Set the alignment of all cells to center both horizontally and vertically
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    else:
        print(index, "===>", 'Failed to retrieve the webpage')
